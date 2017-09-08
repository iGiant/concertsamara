import logging
import urllib.parse
import requests
from lxml import html
from tqdm import trange


def getafisha():
    def addtusa(mystr):
        temp = tuple(parsed_body.xpath(mystr))
        return temp[0] if temp else ''

    def changequotes(mytext):
        if mytext and mytext[0] == '"':
            mytext = '«' + mytext[1:]
        mytext = mytext.replace(' "', ' «')
        return mytext.replace('"', '»')

    def load_setting()-> tuple:
        with open('subscription.dat', 'r', encoding='utf-8') as file:
            subsriptions = file.read().split('\n')
        result = []
        for subsription in subsriptions:
            if subsription:
                mydict = {}
                parametrs = subsription.split(',')
                mydict['keys'] = list(parametrs[:-2])
                mydict['mail'] = parametrs[-2]
                mydict['count'] = parametrs[-1]
                result.append(mydict)
        return tuple(result)

    def search(eventlist: tuple, subscription: tuple):

        def send_mail(mail: str, key: str, event: dict):
            import smtplib
            from concertsamaradata import SMTPSERVER, SERVICESMAILLOGIN, SERVICESMAILPASS, BACKMAILADRR
            from email.mime.text import MIMEText
            mailtext = f"""Сработал триггер на слово "<b>{key}</b>"<br>
Мероприятие <a href="{event['url']}">{event['name']}</a> пройдет {event['date']} ({event['time'][:2]}) 
в {event['time'][3:]} в следующем месте: <i>{event['place']}</i>.<br><br>Билеты можно купить
<a href="{event['buy']}">здесь</a><br><br>{event['detail']}"""
            msg = MIMEText(mailtext, 'HTML', 'utf-8')
            msg['Subject'] = f'Культурные мероприятия Самары: сработал триггер {key}'
            msg['From'] = BACKMAILADRR
            msg['To'] = mail
            smtpObj = smtplib.SMTP(SMTPSERVER, 587)
            smtpObj.ehlo()
            smtpObj.starttls()
            smtpObj.ehlo()
            smtpObj.login(SERVICESMAILLOGIN, SERVICESMAILPASS)
            try:
                smtpObj.sendmail(BACKMAILADRR, mail, msg.as_string())
            except:
                pass
            smtpObj.quit()

        for event in eventlist:
            for index, keys in enumerate(subscription):
                fbreak = False
                for key in keys['keys']:
                    if event['name'].lower().find(key.lower()) != -1 or event['detail'].lower().find(key.lower()) != -1:
                        fbreak = True
                        send_mail(subscription[index]['mail'], key, event)
                        logger.info(f'''Письмо на "{subscription[index]['mail']}", кодовое слово "{key}"''')
                        break
                if fbreak:
                    if keys['count'] != '0':
                        subscription[index]['count'] = '-1' if keys['count'] == '1' else str(int(keys['count']) - 1)
                    break
        with open('subscription.dat', 'w', encoding='utf-8') as file:
            for keys in subscription:
                if keys['count'] != '-1':
                    text = ','.join(keys['keys'])
                    file.write(f"{text},{keys['mail']},{keys['count']}\n")

    logging.basicConfig(format='%(asctime)s - %(message)s', level=logging.INFO, filename=r'd:\python\logs\parsing.log')
    logger = logging.getLogger(__name__)
    logger.info('*' * 25)
    http = 'http://koncertsamara.ru/afisha/'
    afisha = []
    pages = 0
    while True:
        response = requests.get(http + '?a-page=' + str(pages))
        parsed_body = html.fromstring(response.text)
        temp_page = parsed_body.xpath('//*[@id="main"]/div[2]/div[3]/ul/li/a/text()')
        if temp_page[-1] != 'Следующая':
            pages = int(temp_page[-1])
            break
        pages = int(temp_page[-2]) - 1
    for page in trange(pages):
        response = requests.get(http + '?a-page=' + str(page))
        parsed_body = html.fromstring(response.text)
        for i in range(1, len(parsed_body.xpath('///*[@id="main"]/div/div/ul/li/div/div[2]/h3/text()')) + 1):
            tusa = {}
            tusa['name'] = changequotes(addtusa(f'///*[@id="main"]/div/div/ul/li[{i}]/div/div[2]/h3/text()'))
            tusa['date'] = addtusa(f'///*[@id="main"]/div/div/ul/li[{i}]/div/div[1]/span[1]/text()')
            tusa['time'] = addtusa(f'///*[@id="main"]/div/div/ul/li[{i}]/div/div[1]/span[3]/text()')
            tusa['place'] = changequotes(addtusa(f'///*[@id="main"]/div/div/ul/li[{i}]/h4/a/text()'))
            tusa['url'] = addtusa(f'///*[@id="main"]/div/div/ul/li[{i}]/div/div[4]/div/a[1]/@href')
            tusa['buy'] = addtusa(f'///*[@id="main"]/div/div/ul/li[{i}]/div/div[4]/div/a[2]/@href')
            if addtusa(f'///*[@id="main"]/div/div/ul/li[{i}]/div/div[4]/div/a[1]/text()') in ['']:
                tusa['url'] = addtusa(f'///*[@id="main"]/div/div/ul/li[{i}]/div/div[4]/div/a[2]/@href')
                tusa['buy'] = addtusa(f'///*[@id="main"]/div/div/ul/li[{i}]/div/div[4]/div/a[3]/@href')
            tusa['url'] = urllib.parse.urljoin(http, tusa['url'])
            tusa['buy'] = urllib.parse.urljoin(http, tusa['buy'])
            response_detail = requests.get(tusa['url'])
            parsed_body_detail = html.fromstring(response_detail.text)
            tusa['detail'] = ''
            temp_detail = list(parsed_body_detail.xpath('//*[ @ id = "current-description"]/p/text()'))
            if temp_detail:
                for j in range(len(temp_detail)):
                    if len(tusa['detail']) < len(temp_detail[j]):
                        tusa['detail'] = temp_detail[j]
            afisha.append(tusa)
    result = tuple(afisha)
    search(result, load_setting())
    return result


def savetofile(afisha, file='koncert.xlsx'):
    import openpyxl
    from openpyxl.styles import fonts, alignment, Side, Border
    from openpyxl.styles.colors import COLOR_INDEX
    from openpyxl.comments import comments
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Дата', 'Время', 'Событие (клик – подробно)', 'Место проведения (клик – бронирование)'])
    side = Side(style='thin', color=COLOR_INDEX[0])
    dside = Side(style='double', color=COLOR_INDEX[0])
    border = Border(left=side, right=side, top=side, bottom=side)
    hborder = Border(left=side, right=side, top=side, bottom=dside)
    for i in range(len(afisha)):
        ws.append([afisha[i]['date'], afisha[i]['time'],
                   '=HYPERLINK("%s","%s")' % (afisha[i]['url'], afisha[i]['name']),
                   '=HYPERLINK("%s","%s")' % (afisha[i]['buy'], afisha[i]['place'])])
        if len(afisha[i]['detail']) > 10:
            ws['C' + str(i + 2)].comment = comments.Comment(afisha[i]['detail'], '')
        for r in ('A', 'B', 'C', 'D'):
            ws[r + str(i + 2)].border = border
            if r in ('A', 'B'):
                ws[r + str(i + 2)].alignment = alignment.Alignment(horizontal='center')
    for sym in ('A1', 'B1', 'C1', 'D1'):
        ws[sym].font = fonts.Font(size=12, bold=True)
        ws[sym].alignment = alignment.Alignment(horizontal='center')
        ws[sym].border = hborder
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 60
    wb.save(file)


if __name__ == '__main__':
    import sys

    if len(sys.argv) > 1:
        if not sys.argv[1].endswith('.xlsx'):
            sys.argv[1] += '.xlsx'
        savetofile(getafisha(), sys.argv[1])
    else:
        afisha = getafisha()
        for i in range(len(afisha)):
            print(f"{afisha[i]['date']} - {afisha[i]['time']} : {afisha[i]['name']} ({afisha[i]['place']})")
        input()
