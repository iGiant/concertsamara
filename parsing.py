import logging
import urllib.parse
import requests
from lxml import html
from tqdm import trange
from sqlalchemy import create_engine, Column, Integer, String
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker
from mysclient import send_message_to_slack

engine = create_engine('sqlite:///D:/python/db/koncertsamara.sqlite', echo=False)
Base = declarative_base()
Session = sessionmaker(engine)

class Subscription(Base):
    __tablename__ = 'subscription'
    id = Column(Integer, primary_key=True)
    trigger = Column(String(50), nullable=False)
    telegram_id = Column(Integer)
    mail = Column(String(50))
    count = Column(Integer)

    def __repr__(self):
        telegram = f'telegram: {self.telegram_id}, ' if self.telegram_id else ''
        mail = f'mail: {self.mail}, ' if self.mail else ''
        return f'trigger: {self.trigger}, {telegram}{mail}count: {self.count}'

def getafisha()-> tuple:
    def addtusa(mystr: str)-> str:
        temp = tuple(parsed_body.xpath(mystr))
        return temp[0] if temp else ''

    def changequotes(mytext: str)-> str:
        if mytext and mytext[0] == '"':
            mytext = '«' + mytext[1:]
        return mytext.replace(' "', ' «').replace('"', '»').lstrip().rstrip()


    def search(eventlist: tuple):

        def send_mail(mail: str, key: str, event: dict):
            import smtplib
            from concertsamaradata import SMTPSERVER, SERVICESMAILLOGIN, SERVICESMAILPASS, BACKMAILADRR
            from email.mime.text import MIMEText
            mailtext = f"""Сработал триггер на слово "<b>{key}</b>"<br>
Мероприятие <a href="{event['url']}">{event['name']}</a> пройдет {event['date']} ({event['time'][:2]}) 
в {event['time'][3:]} в следующем месте: <i>{event['place']}</i>.<br><br>Билеты можно купить
<a href="{event['buy']}">здесь</a><br><br>{event['detail']}"""
            msg = MIMEText(mailtext, 'HTML', 'utf-8')
            msg['Subject'] = f'Культурные мероприятия Самары: сработал триггер {key}'
            msg['From'] = BACKMAILADRR
            msg['To'] = mail
            smtpObj = smtplib.SMTP(SMTPSERVER, 587)
            smtpObj.ehlo()
            smtpObj.starttls()
            smtpObj.ehlo()
            smtpObj.login(SERVICESMAILLOGIN, SERVICESMAILPASS)
            try:
                smtpObj.sendmail(BACKMAILADRR, mail, msg.as_string())
            except Exception:
                pass
            smtpObj.quit()

        session = Session()
        for event in eventlist:
            if session.query(Subscription).count() > 0:
                for subscr in session.query(Subscription).all():
                    if subscr.count != -1 and (subscr.trigger.lower() in event.get('name', '').lower() or
                                               subscr.trigger.lower() in event.get('detail', '').lower()):
                        if subscr.count != 0:
                            if subscr.count == 1:
                                subscr.count = -1
                            else:
                                subscr.count -= 1
                            session.commit()
                        telegram_text = (f"Сработал триггер на слово *{subscr.trigger}*,\n" +
                                             f"Мероприятие {event['name']} пройдет {event['date']} в {event['place']}")
                        send_message_to_slack(':sound: Concert', telegram_text)
                        if subscr.mail:
                            send_mail(subscr.mail, subscr.trigger, event)
                        logger.info(f'''Письмо на "{subscr.mail}", кодовое слово "{subscr.trigger}"''')
            session.close()

    logging.basicConfig(format='%(asctime)s - %(message)s', level=logging.INFO, filename=r'd:\python\logs\parsing.log')
    logger = logging.getLogger(__name__)
    logger.info('*' * 25)
    http = 'http://koncertsamara.ru/afisha/'
    afisha = []
    pages = 0
    while True:
        temp_response = html.fromstring(requests.get(f'{http}?a-page={pages}').text)
        temp_page = temp_response.xpath('//div[@class="pagination"]/ul/li/a/text()')
        if temp_page[-1] != 'Следующая':
            pages = int(temp_page[-1])
            break
        pages = int(temp_page[-2]) - 1
    for page in trange(pages):
        response = requests.get(http + '?a-page=' + str(page))
        parsed_body = html.fromstring(response.text)
        for i in range(1, round(parsed_body.xpath('count(//ul[@class="list"]/li)')) + 1):
            tusa = {}
            tusa['name'] = changequotes(addtusa(f'//ul[@class="list"]/li[{i}]/div/div[2]/h3/text()'))
            tusa['date'] = addtusa(f'//ul[@class="list"]/li[{i}]/div/div[1]/span[1]/text()')
            tusa['time'] = addtusa(f'//ul[@class="list"]/li[{i}]/div/div[1]/span[3]/text()')
            tusa['place'] = changequotes(addtusa(f'//ul[@class="list"]/li[{i}]/h4/a/text()'))
            tusa['url'] = addtusa(f'//ul[@class="list"]/li[{i}]/div/div[4]/div/a[1]/@href')
            tusa['buy'] = addtusa(f'//ul[@class="list"]/li[{i}]/div/div[4]/div/a[2]/@href')
            if not tusa['url']:
                tusa['url'] = addtusa(f'//ul[@class="list"]/li[{i}]/div/div[4]/div/a[2]/@href')
                tusa['buy'] = addtusa(f'//ul[@class="list"]/li[{i}]/div/div[4]/div/a[3]/@href')
            tusa['url'] = urllib.parse.urljoin(http, tusa['url'])
            tusa['buy'] = urllib.parse.urljoin(http, tusa['buy'])
            temp_response = html.fromstring(requests.get(tusa['url']).text)
            temp_detail = temp_response.xpath('//*[@id="current-description"]/p/text()')
            tusa['detail'] = max(temp_detail, key=len) if temp_detail else ''
            afisha.append(tusa)
    result = tuple(afisha)
    search(result)
    return result


def savetofile(afisha, file='koncert.xlsx'):
    import openpyxl
    from openpyxl.styles import fonts, alignment, Side, Border
    from openpyxl.styles.colors import COLOR_INDEX
    from openpyxl.comments import comments
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Дата', 'Время', 'Событие (клик – подробно)', 'Место проведения (клик – бронирование)'])
    side = Side(style='thin', color=COLOR_INDEX[0])
    dside = Side(style='double', color=COLOR_INDEX[0])
    border = Border(left=side, right=side, top=side, bottom=side)
    hborder = Border(left=side, right=side, top=side, bottom=dside)
    for i in range(len(afisha)):
        ws.append([afisha[i]['date'], afisha[i]['time'],
                   '=HYPERLINK("%s","%s")' % (afisha[i]['url'], afisha[i]['name']),
                   '=HYPERLINK("%s","%s")' % (afisha[i]['buy'], afisha[i]['place'])])
        if len(afisha[i]['detail']) > 10:
            ws['C' + str(i + 2)].comment = comments.Comment(afisha[i]['detail'], '')
        for r in ('A', 'B', 'C', 'D'):
            ws[r + str(i + 2)].border = border
            if r in ('A', 'B'):
                ws[r + str(i + 2)].alignment = alignment.Alignment(horizontal='center')
    for sym in ('A1', 'B1', 'C1', 'D1'):
        ws[sym].font = fonts.Font(size=12, bold=True)
        ws[sym].alignment = alignment.Alignment(horizontal='center')
        ws[sym].border = hborder
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 60
    wb.save(file)


if __name__ == '__main__':
    import sys

    if len(sys.argv) > 1:
        if not sys.argv[1].endswith('.xlsx'):
            sys.argv[1] += '.xlsx'
        savetofile(getafisha(), sys.argv[1])
    else:
        afisha = getafisha()
        print(*[f"{event['date']} - {event['time']} : {event['name']} ({event['place']})" for event in afisha],
              sep = '\n')
        input()
