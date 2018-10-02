import asyncio
import urllib.parse
import sys
from aiohttp import ClientSession
from lxml import html
from fake_useragent import UserAgent
import openpyxl
from openpyxl.styles import fonts, alignment, Side, Border
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.comments import comments
from tqdm import tqdm
from dataclasses import dataclass


@dataclass
class Event:
    name = ''
    date = ''
    time = ''
    place = ''
    url = ''
    buy = ''
    detail = ''


addr = 'http://koncertsamara.ru/afisha/'


async def get_source(session, number, url, headers, bar):
    async with session.get(url, headers=headers) as response:
        if bar is not None:
            bar.update(1)
        return number, await response.read()


async def get_last_page(session, pages, headers)-> int:
    while True:
        url = f"{addr}?a-page={pages}"
        source = (await asyncio.gather(asyncio.ensure_future(get_source(session, 0, url, headers, None))))[0]
        http = html.fromstring(source[1])
        last_page = http.xpath('//div[@class="pagination"]/ul/li/a/text()')
        if last_page[-1] == 'Следующая':
            pages = int(last_page[-2]) - 1
        else:
            return int(last_page[-1])


def changequotes(mytext: str) -> str:
    if mytext and mytext[0] == '"':
        mytext = '«' + mytext[1:]
    return mytext.replace(' "', ' «').replace('"', '»').strip()


def get_element(source: html, path: str)-> str:
    result = source.xpath(path)
    return result[0] if result else ''


async def main():
    ua = UserAgent()
    headers = {'User-Agent': ua.ie}
    result = []
    index = 0
    async with ClientSession() as session:
        pages = await get_last_page(session, 0, headers)
        bar = tqdm(total=pages, desc='Обработка страниц')
        tasks = [asyncio.ensure_future(
            get_source(session, page, f"{addr}?a-page={page}", headers, bar)) for page in range(pages)]
        responses = await asyncio.gather(*tasks)
        bar.close()
        responses.sort()
        tasks = []
        bar = tqdm(desc='Загрузка мероприятий')
        for response in responses:
            source = html.fromstring(response[1])
            for i in range(1, round(source.xpath('count(//ul[@class="list"]/li)')) + 1):
                event = Event()
                event.name = changequotes(get_element(source, f'//ul[@class="list"]/li[{i}]/div/div[2]/h3/text()'))
                event.date = get_element(source, f'//ul[@class="list"]/li[{i}]/div/div[1]/span[1]/text()')
                event.time = get_element(source, f'//ul[@class="list"]/li[{i}]/div/div[1]/span[3]/text()')
                event.place = changequotes(get_element(source, f'//ul[@class="list"]/li[{i}]/h4/a/text()'))
                event.url = get_element(source, f'//ul[@class="list"]/li[{i}]/div/div[4]/div/a[1]/@href')
                event.buy = get_element(source, f'//ul[@class="list"]/li[{i}]/div/div[4]/div/a[2]/@href')
                if not event.url or event.url == '/newslist/novinka-elektronnyj-bilet/':
                    event.url = get_element(source, f'//ul[@class="list"]/li[{i}]/div/div[4]/div/a[2]/@href')
                    event.buy = get_element(source, f'//ul[@class="list"]/li[{i}]/div/div[4]/div/a[3]/@href')
                event.url = urllib.parse.urljoin(addr, event.url)
                event.buy = urllib.parse.urljoin(addr, event.buy)
                tasks.append(asyncio.ensure_future(
                    get_source(session, index, event.url, headers, bar)))
                index += 1
                result.append(event)
        bar.total = len(result)
        resp = await asyncio.gather(*tasks)
        bar.close()

        resp.sort()
        for i, item in enumerate(result):
            temp_resp = html.fromstring(resp[i][1])
            temp_detail = temp_resp.xpath('//*[@id="current-description"]/p/text()')
            item.detail = max(temp_detail, key=len) if temp_detail else ''
        return result


def savetofile(afisha_, file='koncert.xlsx'):

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Дата', 'Время', 'Событие (клик – подробно)', 'Место проведения (клик – бронирование)'])
    side = Side(style='thin', color=COLOR_INDEX[0])
    dside = Side(style='double', color=COLOR_INDEX[0])
    border = Border(left=side, right=side, top=side, bottom=side)
    hborder = Border(left=side, right=side, top=side, bottom=dside)
    for i in range(len(afisha_)):
        ws.append([afisha_[i].date, afisha_[i].time,
                   '=HYPERLINK("%s","%s")' % (afisha_[i].url, afisha_[i].name),
                   '=HYPERLINK("%s","%s")' % (afisha_[i].buy, afisha_[i].place)])
        if len(afisha_[i].detail) > 10:
            ws['C' + str(i + 2)].comment = comments.Comment(afisha_[i].detail, '')
        for r in ('A', 'B', 'C', 'D'):
            ws[r + str(i + 2)].border = border
            if r in ('A', 'B'):
                ws[r + str(i + 2)].alignment = alignment.Alignment(horizontal='center')
    for sym in ('A1', 'B1', 'C1', 'D1'):
        ws[sym].font = fonts.Font(size=12, bold=True)
        ws[sym].alignment = alignment.Alignment(horizontal='center')
        ws[sym].border = hborder
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 60
    wb.save(file)


if __name__ == '__main__':
    afisha = asyncio.run(main())
    if len(sys.argv) > 1:
        if not sys.argv[1].endswith('.xlsx'):
            sys.argv[1] += '.xlsx'
        savetofile(afisha, sys.argv[1])
    else:
        print()
        print(*[f"{event.date} - {event.time} : {event.name} ({event.place})" for event in afisha],
              sep='\n')
